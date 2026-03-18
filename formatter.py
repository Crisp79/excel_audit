# Copyright (c) 2026 Akshith Vuppala
# Licensed under the MIT License.
# See LICENSE file in the project root for full license information.

import openpyxl
from openpyxl.styles import PatternFill


def format_and_split_dataframe(file_path, df, output_f):
    wb = openpyxl.load_workbook(file_path)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Group by sheet so we only determine the 'Variation' column once per sheet
    for sheet_name, group in df.groupby("SHEET"):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Determine the Variation column index ONCE per sheet
        # We check if a 'VARIATION' header already exists; if not, use max_column + 1
        var_col_idx = None
        for cell in ws[1]:  # Look at the header row
            if cell.value == "VARIATION":
                var_col_idx = cell.column
                break

        if var_col_idx is None:
            var_col_idx = ws.max_column + 1
            ws.cell(row=1, column=var_col_idx).value = "VARIATION"

        # Now iterate through the rows for this specific sheet
        for _, data in group.iterrows():
            row_idx = int(data["ROW_NUM"])

            # 1. Write the variation value in the fixed column
            ws.cell(row=row_idx, column=var_col_idx).value = data["VARIATION"]

            # 2. Apply Audit Coloring Logic
            m_val = data["_merge"]
            variation = data["VARIATION"]
            debit_a = data.get("DEBIT_A", data.get("DEBIT", 0))
            credit_b = data.get("CREDIT_B", data.get("CREDIT", 0))

            row_fill = None
            if m_val == "both":
                row_fill = green if variation == 0 else yellow
            elif m_val == "left_only" and debit_a != 0:
                row_fill = red
            elif m_val == "right_only" and credit_b != 0:
                row_fill = red

            # 3. Apply the fill to the entire row
            if row_fill:
                for cell in ws[row_idx]:
                    cell.fill = row_fill

    wb.save(output_f)
    print(f"Original file {output_f} updated successfully.")
