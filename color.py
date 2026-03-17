import openpyxl
from openpyxl.styles import PatternFill


def format_and_split_audit_excel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active  # Assuming the raw data is in the first sheet

    # 1. Define the colors
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 2. Map headers to column indices
    headers = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    header_list = [cell.value for cell in ws[1]]  # To copy headers to new sheets

    def find_col(possible_names):
        for name in possible_names:
            if name in headers:
                return headers[name]
        return None

    # Identify indices
    m_idx = find_col(["_merge"])
    v_idx = find_col(["VARIATION"])
    d_a_idx = find_col(["DEBIT_a", "DEBIT"])
    c_b_idx = find_col(["CREDIT_b", "CREDIT"])
    sheet_col_idx = find_col(["SHEET"])  # Identify the SHEET column

    if not sheet_col_idx:
        print("Error: 'SHEET' column not found.")
        return

    # 3. Iterate through rows, color them, and prepare to move them
    # We store rows in a dictionary: {sheet_name: [list_of_rows]}
    sheets_data = {}

    for row in ws.iter_rows(min_row=2):
        m_val = row[m_idx - 1].value if m_idx else None
        variation = row[v_idx - 1].value if v_idx else 0
        debit_a = row[d_a_idx - 1].value if d_a_idx else 0
        credit_b = row[c_b_idx - 1].value if c_b_idx else 0

        # Determine the destination sheet name
        sheet_name = (
            str(row[sheet_col_idx - 1].value)
            if row[sheet_col_idx - 1].value
            else "Other"
        )
        # Clean sheet name (Excel limits)
        sheet_name = "".join(c for c in sheet_name if c not in r"\*?:/[]")[:31]

        # Apply Your Existing Logic
        row_fill = None
        if m_val == "both":
            row_fill = green if variation == 0 else yellow
        elif m_val == "left_only" and debit_a and debit_a != 0:
            row_fill = red
        elif m_val == "right_only" and credit_b and credit_b != 0:
            row_fill = red

        if row_fill:
            for cell in row:
                cell.fill = row_fill

        # Add row to our dictionary
        if sheet_name not in sheets_data:
            sheets_data[sheet_name] = []
        sheets_data[sheet_name].append([cell.value for cell in row])

        # Carry over the formatting - openpyxl requires cell-by-cell copy for styles
        # So we store the fill as well
        sheets_data[sheet_name][-1] = (sheets_data[sheet_name][-1], row_fill)

    # 4. Create new sheets and write data
    for sheet_name, rows in sheets_data.items():
        new_ws = wb.create_sheet(title=sheet_name)
        new_ws.append(header_list)

        for row_data, fill in rows:
            new_ws.append(row_data)
            if fill:
                for cell in new_ws[new_ws.max_row]:
                    cell.fill = fill

        # 5. Column Cleanup
        # We handle indices carefully. If we delete m_idx first and it's
        # to the left of sheet_col_idx, the sheet index would shift.

        cols_to_delete = []
        if m_idx:
            cols_to_delete.append(m_idx)
        if sheet_col_idx:
            cols_to_delete.append(sheet_col_idx)

        # Sort in reverse order so deleting one doesn't shift the index of the other
        for col_idx in sorted(cols_to_delete, reverse=True):
            new_ws.delete_cols(col_idx)

        new_ws.freeze_panes = "A2"

    # 5. Final Cleanup: Remove the original source sheet
    wb.remove(ws)

    wb.save(filename)
    print(f"File successfully formatted and split into {len(sheets_data)} sheets.")


# Usage
# format_and_split_audit_excel("audit_report.xlsx")
